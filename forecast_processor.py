import os 
import json
import pandas as pd
from pathlib import Path
from typing import List, Dict, Any, Optional, Union
import openpyxl
from app.services.forecast import forecast_table
from app.core.logger import setup_logger


BASE_OUTPUT_DIR = Path('output')
LLM_OUTPUT_DIR = BASE_OUTPUT_DIR / "LLM output"
PROCESSED_OUTPUT_DIR = BASE_OUTPUT_DIR / "processed"

logger = setup_logger()

def load_po_data(file_path: Path) -> Optional[List[Dict[str, Any]]]:
    """Load purchase order data from a JSON file."""
    
    try:
        with open(file_path, 'r', encoding='utf-8') as f: # Added encoding
            po_data_list = json.load(f)
        logger.info(f"JSON data loaded successfully from {file_path}.")
        return po_data_list
    except FileNotFoundError:
        logger.error(f"The file at path {file_path} was not found. Cannot perform forecasting.")
    except json.JSONDecodeError as e: # Catch specific error
        logger.error(f"Could not decode JSON from {file_path}. Is it valid JSON? Error: {e}. Cannot perform forecasting.")
    except Exception as e:
        logger.error(f"An unexpected error occurred during loading PO data: {e}") # Clarified error message
    return None

def process_forecast_data(po_data_list: List[Dict[str, Any]]) -> pd.DataFrame:
    """Process PO data into a consolidated forecast DataFrame."""
    all_forecast_dfs: List[pd.DataFrame] = []
    for po_data in po_data_list:
        if isinstance(po_data, dict):
            try: # Added try-except for forecast_table call
                df = forecast_table(po_data)
                all_forecast_dfs.append(df)
            except Exception as e:
                logger.error(f"Error processing PO data for forecasting: {po_data.get('po_id', 'N/A')}. Error: {e}. Skipping this PO.")
        else:
            logger.warning(f"Warning: Expected dictionary for PO data but got {type(po_data)}. Skipping entry.")
    if not all_forecast_dfs:
        logger.warning("No valid PO data to process into forecast. Returning empty DataFrame.")
        return pd.DataFrame()
    combined_df = pd.concat(all_forecast_dfs, ignore_index=True)
    return combined_df

def clean_forecast_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """Clean and standardize the forecast DataFrame."""
    if df.empty:
        logger.warning("Input DataFrame for cleaning is empty. Returning as is.")
        return df

    # Round "Inflow (USD)" to 2 decimal places
    if "Inflow (USD)" in df.columns:
        df["Inflow (USD)"] = pd.to_numeric(df["Inflow (USD)"], errors='coerce')
        df["Inflow (USD)"] = df["Inflow (USD)"].fillna(0.0).round(2) # Chain fillna and round

    # Standardize 'Month' column to 'YYYY-MM' string
    if 'Month' in df.columns:
        try:
            
            df['Month'] = df['Month'].astype(str)
            datetime_months = pd.to_datetime(df['Month'], errors='coerce')
            df['Month'] = datetime_months.dt.strftime('%Y-%m')
            df.dropna(subset=['Month'], inplace=True)
        except Exception as e:
            logger.warning(f"Could not standardize 'Month' column format: {e}. Proceeding with existing format.")
            # Ensure it's still string if conversion failed
            if not pd.api.types.is_string_dtype(df['Month']):
                df['Month'] = df['Month'].astype(str)

    if "Client Name" in df.columns:
        df["Client Name"] = df["Client Name"].astype(str).str.strip()
    if "PO No" in df.columns:
        df["PO No"] = df["PO No"].astype(str).str.strip()
    if "Project Owner" in df.columns: # Added this if it's a critical column
        df["Project Owner"] = df["Project Owner"].astype(str).str.strip()


    # Remove duplicate rows after standardization
    df.drop_duplicates(inplace=True, keep='last')
    logger.info("Forecast DataFrame cleaned successfully.")
    return df

# --- Changed output_csv_path type hint to Path ---
def save_forecast_csv(df: pd.DataFrame, output_csv_path: Path) -> None:
    """Save the forecast DataFrame to CSV."""
    # Ensure directory exists before saving
    output_csv_path.parent.mkdir(parents=True, exist_ok=True)
    df.to_csv(output_csv_path, index=False, float_format='%.2f')
    logger.info(f"Saved forecast table to '{output_csv_path}'")

# --- Changed pivot_excel_path type hint to Path ---
def generate_and_save_pivot(df: pd.DataFrame, pivot_excel_path: Path) -> None:
    """Generate a pivot table from the forecast DataFrame and save to Excel."""
    if df.empty or 'Month' not in df.columns or df['Month'].isnull().all():
        logger.warning("No valid data for pivot table generation. Skipping pivot.")
        return
    try:
        # Ensure 'Month' column is treated as string for consistent parsing
        df['Month'] = df['Month'].astype(str)
        month_as_datetime = pd.to_datetime(df['Month'], format='%Y-%m', errors='coerce')
        month_as_datetime.dropna(inplace=True)
        if month_as_datetime.empty:
            logger.error("No valid 'Month' data in 'YYYY-MM' format for pivot table. Skipping pivot generation.")
            return

        min_month = month_as_datetime.min()
        max_month = month_as_datetime.max()
        # Ensure min_month and max_month are not NaT before range
        if pd.isna(min_month) or pd.isna(max_month):
            logger.error("Min or Max month is NaT after conversion. Cannot generate month range for pivot. Skipping.")
            return

        all_months_for_pivot = pd.date_range(min_month, max_month, freq='MS').strftime('%Y-%m')

        pivot = df.pivot_table(
            index=["Client Name", "PO No", "Project Owner", "Status"],
            columns="Month",
            values="Inflow (USD)",
            aggfunc="sum",
            fill_value=0.0
        )

        # Ensure all months are present as columns
        # Filter all_months_for_pivot to only include those that are strings
        valid_month_cols = [col for col in all_months_for_pivot if isinstance(col, str)]
        pivot = pivot.reindex(columns=valid_month_cols, fill_value=0.0)
        pivot.reset_index(inplace=True)

        if not pivot.empty:
            # Check for 'S.No' *before* dropping it if it's not meant to be there
            # Your original code has a logical inconsistency here: it drops 'S.No'
            # then adds it again. If 'S.No' is never meant to be an input column
            # but always a generated output column, the drop is unnecessary.
            # Assuming 'S.No' is always generated:
            if 'S.No' in pivot.columns: # Remove if it was somehow in source data
                pivot = pivot.drop(columns=['S.No'])

            # Generate S.No
            pivot.insert(0, 'S.No', range(1, 1 + len(pivot)))

            # Reorder columns - make sure all expected index columns are present
            month_cols = [col for col in pivot.columns if col not in ['S.No', 'Client Name', 'PO No', 'Project Owner', 'Status']]
            # Ensure 'Client Name', 'PO No', 'Project Owner', 'Status' exist before creating new_order
            expected_index_cols = ['Client Name', 'PO No', 'Project Owner', 'Status']
            current_index_cols = [col for col in expected_index_cols if col in pivot.columns]
            new_order = ['S.No'] + current_index_cols + sorted(month_cols) # Sort month columns for consistency

            pivot = pivot[new_order]


        # Round float columns to integer for Excel output
        for col in pivot.columns:
            # Only apply if it's a numeric column that contains floats
            if pd.api.types.is_numeric_dtype(pivot[col]) and not pd.api.types.is_integer_dtype(pivot[col]):
                pivot[col] = pivot[col].round(0).astype(int)

        # Ensure directory exists before saving
        pivot_excel_path.parent.mkdir(parents=True, exist_ok=True)
        with pd.ExcelWriter(pivot_excel_path, engine='openpyxl') as writer:
            pivot.to_excel(writer, index=False, sheet_name="Forecast")
            worksheet = writer.sheets["Forecast"]

            # Dynamically determine the start column for month values
            fixed_cols = ['S.No', 'Client Name', 'PO No', 'Project Owner', 'Status']
            # Find the index of the first month column in the *actual* pivot DataFrame columns
            # This is safer than fixed offsets.
            try:
                first_month_col_name = sorted(month_cols)[0] if month_cols else None
                if first_month_col_name:
                    first_month_col_idx = pivot.columns.get_loc(first_month_col_name) + 1 # +1 for 1-based indexing
                else:
                    first_month_col_idx = worksheet.max_column + 1 # No month columns, skip formatting
            except IndexError: # If month_cols is empty
                first_month_col_idx = worksheet.max_column + 1 # No month columns, skip formatting

            if first_month_col_idx <= worksheet.max_column: # Only proceed if there are month columns
                # Apply number formatting to month columns in Excel
                for row in worksheet.iter_rows(min_row=2, # Start from second row (after header)
                                               min_col=first_month_col_idx, # Use calculated column index
                                               max_col=worksheet.max_column):
                    for cell in row:
                        if cell.value is not None:
                            try:
                                # Check if the value can be converted to float before formatting
                                float_val = float(cell.value)
                                cell.number_format = '0' # Format as integer
                            except (ValueError, TypeError): # Catch TypeError for non-numeric types
                                pass

        logger.info(f"Saved forecast pivot to '{pivot_excel_path}'")
    except Exception as e:
        logger.error(f"Error during pivot table generation: {e}. Skipping pivot generation.")

def run_forecast_processing(
    # --- CRITICAL FIX: Change type hints and default values to Path objects ---
    input_json_path: Path = LLM_OUTPUT_DIR / "purchase_orders.json",
    output_csv_path: Path = PROCESSED_OUTPUT_DIR / "forecast_output.csv",
    pivot_excel_path: Path = PROCESSED_OUTPUT_DIR / "forecast_pivot.xlsx"
) -> None:
    """
    Loads PO data from JSON, generates a forecast table,
    and saves it to CSV and a pivot table to Excel.

    Args:
        input_json_path (Path): Path to the input JSON file containing PO data.
        output_csv_path (Path): Path to the output CSV file for the forecast table.
        pivot_excel_path (Path): Path to the output Excel file for the pivot table.
    """
    logger.info("Running Forecast Processor")
    # No need for file_path = Path(input_json_path); input_json_path is already a Path object
    po_data_list = load_po_data(input_json_path) # Pass Path object directly to load_po_data
    if po_data_list is None:
        logger.error("Failed to load PO data. Aborting forecast processing.")
        return

    combined_df = process_forecast_data(po_data_list)
    combined_df = clean_forecast_dataframe(combined_df)

    if combined_df.empty:
        logger.warning("No forecast data available to process. Skipping CSV and Excel output.")
        return

    # Pass Path objects directly to save functions
    save_forecast_csv(combined_df, output_csv_path)
    generate_and_save_pivot(combined_df, pivot_excel_path)